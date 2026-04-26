"""Excel BOQ adapter — lazy `openpyxl` import + thin core delegation.

`parse_boq_xlsx(content)` reads the first worksheet, finds a header
row (anywhere in the first 20 rows — users sometimes prepend a title /
project info), and emits one `BoqRow` per body row that has a
non-empty description.

`render_boq_xlsx(rows)` writes a single-sheet workbook with the column
order from `core.rows_to_grid`. Number cells get `#,##0` formatting so
totals display with thousand separators; the description column gets
wrapped text so long lines don't blow out the column.

Both functions take/return `bytes` so they're easy to plug into a
FastAPI `UploadFile` / `StreamingResponse` without any framework
imports here.
"""

from __future__ import annotations

import io
import logging

from .core import BoqIOError, BoqRow, coerce_decimal, coerce_str, detect_columns, rows_to_grid

logger = logging.getLogger(__name__)


# How far down the sheet to scan for a header row before giving up. 20
# is generous enough for a project-info banner; beyond that the file
# probably isn't a BOQ.
_HEADER_SCAN_LIMIT = 20


def parse_boq_xlsx(content: bytes) -> list[BoqRow]:
    """Parse a BOQ from raw .xlsx bytes. Raises `BoqIOError` on failure."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover — deployed env always has it
        raise BoqIOError("openpyxl not installed; cannot parse .xlsx") from exc

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise BoqIOError(f"could not open .xlsx content: {exc}") from exc

    sheet = wb.active
    if sheet is None:
        raise BoqIOError("workbook has no active sheet")

    # Read all rows once. xlsx is in-memory at this point so a second
    # iteration is free; the explicit list also lets us index into it.
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx, cols = None, None
    for i, row in enumerate(rows[:_HEADER_SCAN_LIMIT]):
        candidate = detect_columns(list(row))
        if candidate is not None:
            header_idx = i
            cols = candidate
            break

    if cols is None or header_idx is None:
        raise BoqIOError(
            "no recognisable BOQ header row found in the first "
            f"{_HEADER_SCAN_LIMIT} rows. Expected a column whose name contains "
            "'mô tả' / 'description' / 'tên công việc'."
        )

    out: list[BoqRow] = []
    for raw in rows[header_idx + 1 :]:
        if raw is None:
            continue
        row = list(raw)
        # Pad short rows so column accesses never IndexError.
        max_idx = max(
            i
            for i in (
                cols.description,
                cols.code,
                cols.unit,
                cols.quantity,
                cols.unit_price_vnd,
                cols.total_price_vnd,
                cols.material_code,
            )
            if i is not None
        )
        if len(row) <= max_idx:
            row = row + [None] * (max_idx + 1 - len(row))

        description = coerce_str(row[cols.description])
        if not description:
            # Skip blank rows AND rows that have only a code or only a
            # number — they're typically subtotal lines or section
            # banners that shouldn't survive the import.
            continue

        out.append(
            BoqRow(
                description=description,
                code=coerce_str(row[cols.code]) if cols.code is not None else None,
                unit=coerce_str(row[cols.unit]) if cols.unit is not None else None,
                quantity=coerce_decimal(row[cols.quantity]) if cols.quantity is not None else None,
                unit_price_vnd=(coerce_decimal(row[cols.unit_price_vnd]) if cols.unit_price_vnd is not None else None),
                total_price_vnd=(
                    coerce_decimal(row[cols.total_price_vnd]) if cols.total_price_vnd is not None else None
                ),
                material_code=(coerce_str(row[cols.material_code]) if cols.material_code is not None else None),
                sort_order=len(out),
            )
        )

    logger.info(
        "boq_io.parse_xlsx: extracted %d rows from header row %d",
        len(out),
        header_idx,
    )
    return out


def render_boq_xlsx(rows: list[BoqRow], *, sheet_name: str = "BOQ") -> bytes:
    """Render `rows` to .xlsx bytes ready for `StreamingResponse`."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise BoqIOError("openpyxl not installed; cannot render .xlsx") from exc

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    header_cells, body_cells = rows_to_grid(rows)

    # Header row: bold + light-grey fill, frozen for scrolling.
    sheet.append(header_cells)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="EEEEEE")
    for col_idx in range(1, len(header_cells) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = bold
        cell.fill = fill
    sheet.freeze_panes = "A2"

    # Body rows.
    for body_row in body_cells:
        sheet.append(body_row)

    # Number formatting for quantity / unit price / total. Column
    # indices match `_EXPORT_HEADERS` in core.py.
    qty_col = 4
    price_cols = (5, 6)
    for r in range(2, sheet.max_row + 1):
        sheet.cell(row=r, column=qty_col).number_format = "#,##0.##"
        for c in price_cols:
            sheet.cell(row=r, column=c).number_format = "#,##0"

    # Description column wraps for readability. Column "B" = description.
    desc_col_letter = "B"
    sheet.column_dimensions[desc_col_letter].width = 50
    for r in range(2, sheet.max_row + 1):
        sheet.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # Sensible default widths for the rest.
    for letter, width in (("A", 8), ("C", 8), ("D", 12), ("E", 16), ("F", 18), ("G", 16)):
        sheet.column_dimensions[letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
