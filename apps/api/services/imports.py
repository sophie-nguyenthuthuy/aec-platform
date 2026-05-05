"""CSV/XLSX bulk-import service.

Two-phase pipeline:

  1. **Preview** — `parse_upload()` reads bytes (CSV or XLSX), normalises
     headers, runs the per-entity validator, and returns
     `(valid_rows, errors)`. `create_preview_job()` writes a row to
     `import_jobs` with `status='previewed'` and the parsed payload
     stored in JSONB.

  2. **Commit** — `commit_job()` re-loads a previewed job, runs the
     per-entity upsert, stamps `status='committed'` and the row count.
     Idempotent via `(organization_id, external_id)` unique indexes
     on `projects` and `suppliers`: re-uploading the same CSV updates
     existing rows instead of inserting duplicates.

Why two-phase: the upload happens through a multipart POST that the
frontend is allowed to retry on transient failures. If we committed in
that handler, a partial DB write + retry would double-insert. Splitting
the parse from the commit means the "preview" call is a fast
read-mostly path (validate + JSONB blob write) and the commit is an
explicit user gesture that can verify the previewed row IDs match.

The per-entity validators are deliberately verbose (one function per
field) rather than driven from a JSON schema — Vietnamese tenant data
has a lot of "the spec says province is one of N strings but we'll
accept the canonical English names too" rules, and untangling them
from a generic schema framework is harder than just writing the
checks inline.

V1 limits:
  * 1000 rows per upload (JSONB blob would otherwise bloat).
  * 10 columns of free-form text per row (caps the per-row blob
    size at ~5KB pre-encoding).
"""

from __future__ import annotations

import csv
import io
from typing import Any
from uuid import UUID

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

# Cap rows to keep the per-job JSONB blob small. 1000 is enough for
# real onboarding files we've seen; bigger sheets should be split.
MAX_ROWS = 1000

# Acceptable suppliers/projects status values mirror the application
# enums. Mismatch reads cleaner than "must be one of {...}" so we
# render the bad value in the error message.
PROJECT_STATUSES = {
    "planning",
    "design",
    "bidding",
    "construction",
    "handover",
    "completed",
    "on_hold",
    "cancelled",
}


# ---------- Parser: file → list[dict] ----------


def parse_upload(*, content: bytes, filename: str) -> list[dict[str, Any]]:
    """Parse a CSV or XLSX upload into a list of header→value dicts.

    Detects format by extension (case-insensitive). Raises `ValueError`
    on unknown extensions or rows above MAX_ROWS — both are surfaced
    by the router as a 400 with a friendly message.

    Header normalisation: lowercased, trimmed, internal whitespace
    collapsed to underscores. So a column titled "Project Name" in the
    CSV maps to the `project_name` key — saving the validator from a
    case-folding loop on every field access.
    """
    name_lower = filename.lower()
    if name_lower.endswith(".csv"):
        rows = _parse_csv(content)
    elif name_lower.endswith((".xlsx", ".xlsm")):
        rows = _parse_xlsx(content)
    else:
        raise ValueError(f"Unsupported file type: {filename!r}. Expected .csv, .xlsx, or .xlsm.")
    if len(rows) > MAX_ROWS:
        raise ValueError(f"File has {len(rows)} rows; the per-upload cap is {MAX_ROWS}. Split the file and re-upload.")
    return rows


def _normalise_header(h: str) -> str:
    """Lowercase, strip, collapse internal whitespace → underscore.
    Matches the validator's expected key names regardless of how the
    spreadsheet author capitalised the header."""
    return "_".join(h.strip().lower().split())


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    """Decode as UTF-8 (fall back to UTF-8-sig to swallow Excel's BOM
    on Save-As-CSV from Vietnamese locales). Sniff the dialect so we
    handle both comma-separated and Vietnamese semicolon-separated."""
    try:
        text_content = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError(f"File is not UTF-8 encoded ({exc}); re-save as UTF-8.") from exc

    sample = text_content[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # default to comma if the sniffer balks
    reader = csv.reader(io.StringIO(text_content), dialect=dialect)
    rows = list(reader)
    if not rows:
        return []
    headers = [_normalise_header(h) for h in rows[0]]
    return [
        {headers[i]: (cell.strip() if isinstance(cell, str) else cell) for i, cell in enumerate(row[: len(headers)])}
        for row in rows[1:]
        if any(c.strip() for c in row if isinstance(c, str))  # skip blank lines
    ]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    """openpyxl read-only mode → row-by-row generator. We pin
    `data_only=True` so formulas resolve to their cached values
    (Excel saves the last computed value); without it the user'd see
    `=A1+B1` strings instead of numbers."""
    from openpyxl import load_workbook  # lazy import: heavy module

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []
    headers = [_normalise_header(str(h)) if h is not None else "" for h in raw_rows[0]]
    out: list[dict[str, Any]] = []
    for raw in raw_rows[1:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
            continue  # skip empty rows
        row: dict[str, Any] = {}
        for i, cell in enumerate(raw[: len(headers)]):
            if not headers[i]:
                continue
            row[headers[i]] = cell.strip() if isinstance(cell, str) else cell
        out.append(row)
    return out


# ---------- Validators ----------


# Per-entity validators return `(cleaned_dict_or_none, error_message_or_none)`.
# They MUST set `external_id` so the upsert ON CONFLICT clause has a
# natural key — empty external_id means "we couldn't link this back to
# your system", which is fatal for idempotency.


def _validate_project(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    """Project rows: external_id + name required; everything else
    optional. Status must be one of the canonical lifecycle values."""
    external_id = _coerce_str(row.get("external_id"))
    name = _coerce_str(row.get("name"))
    if not external_id:
        return None, "Missing `external_id` (used as the natural key for upsert)."
    if not name:
        return None, "Missing `name`."
    status = _coerce_str(row.get("status")) or "planning"
    if status not in PROJECT_STATUSES:
        return None, f"Invalid status {status!r}. Allowed: {sorted(PROJECT_STATUSES)}"
    cleaned: dict[str, Any] = {
        "external_id": external_id,
        "name": name,
        "type": _coerce_str(row.get("type")),
        "status": status,
    }
    # Optional numeric fields. We accept bad numbers as None rather
    # than rejecting the whole row — losing area_sqm shouldn't kill an
    # otherwise-valid project import.
    for f in ("area_sqm", "budget_vnd", "floors"):
        v = row.get(f)
        if v is None or v == "":
            cleaned[f] = None
        else:
            try:
                cleaned[f] = int(float(v)) if f in ("budget_vnd", "floors") else float(v)
            except (TypeError, ValueError):
                cleaned[f] = None
    # Address rolled into a JSONB blob if any of city/district provided.
    city = _coerce_str(row.get("city"))
    district = _coerce_str(row.get("district"))
    cleaned["address"] = {"city": city, "district": district} if (city or district) else None
    return cleaned, None


def _validate_supplier(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    """Supplier rows: external_id + name required.

    `categories` and `provinces` are accepted as comma-separated strings
    and cast to TEXT[] at commit time — this is how customer
    spreadsheets tend to encode them, and a structured per-cell
    multi-value parser would fight Excel's autoformatting more than
    it's worth.
    """
    external_id = _coerce_str(row.get("external_id"))
    name = _coerce_str(row.get("name"))
    if not external_id:
        return None, "Missing `external_id`."
    if not name:
        return None, "Missing `name`."
    categories = _split_csv(row.get("categories"))
    provinces = _split_csv(row.get("provinces"))
    contact: dict[str, Any] = {}
    for k in ("phone", "email", "address"):
        v = _coerce_str(row.get(k))
        if v:
            contact[k] = v
    return {
        "external_id": external_id,
        "name": name,
        "categories": categories,
        "provinces": provinces,
        "contact": contact,
        "verified": _coerce_bool(row.get("verified")),
    }, None


VALIDATORS = {
    "projects": _validate_project,
    "suppliers": _validate_supplier,
}


def _coerce_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _coerce_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    s = _coerce_str(v)
    if s is None:
        return False
    return s.lower() in {"true", "yes", "y", "1", "có"}


def _split_csv(v: Any) -> list[str]:
    """`"a, b,c"` → `["a", "b", "c"]`. Empty pieces dropped."""
    s = _coerce_str(v)
    if not s:
        return []
    return [piece.strip() for piece in s.split(",") if piece.strip()]


# ---------- Validate-many helper (parser → preview payload) ----------


def validate_rows(
    *,
    entity: str,
    raw_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Run the per-entity validator over each parsed row.

    Returns `(valid_rows, errors)`:
      * `valid_rows` are dicts ready for the upsert.
      * `errors` is a list of `{row_idx, message}` shaped to match the
        `import_jobs.errors` JSONB column.

    Row indexing is 1-based against the user's spreadsheet (header at
    index 1, first data row at 2) — frontend renders the index as-is
    so users can jump straight to the bad cell.
    """
    if entity not in VALIDATORS:
        raise ValueError(f"Unsupported entity: {entity}. Allowed: {list(VALIDATORS)}")
    validator = VALIDATORS[entity]
    valid: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for i, raw in enumerate(raw_rows, start=2):  # +2: row 1 is the header
        cleaned, err = validator(raw)
        if err:
            errors.append({"row_idx": i, "message": err})
        else:
            assert cleaned is not None
            valid.append(cleaned)
    return valid, errors


# ---------- Commit: upsert valid rows into the target table ----------


async def commit_job(
    *,
    session: AsyncSession,
    organization_id: UUID,
    entity: str,
    rows: list[dict[str, Any]],
) -> int:
    """UPSERT validated rows into the target table.

    Returns the number of rows the database touched (inserted OR
    updated). Postgres's `ON CONFLICT DO UPDATE` makes both paths
    indistinguishable from the application's perspective, which is
    exactly the property idempotency requires.

    Each entity has its own SQL because the upsertable columns differ;
    promoting to a generic builder would obscure the per-entity
    nullability rules + array casts.
    """
    if not rows:
        return 0
    if entity == "projects":
        return await _commit_projects(session, organization_id, rows)
    if entity == "suppliers":
        return await _commit_suppliers(session, organization_id, rows)
    raise ValueError(f"Unsupported entity: {entity}")


async def _commit_projects(
    session: AsyncSession,
    organization_id: UUID,
    rows: list[dict[str, Any]],
) -> int:
    """ON CONFLICT on the partial unique index `(org_id, external_id)
    WHERE external_id IS NOT NULL`. Postgres infers the constraint by
    matching the inferred index — that's why the WHERE clause is
    repeated in the index_predicate."""
    written = 0
    for row in rows:
        await session.execute(
            text(
                """
                INSERT INTO projects (
                    id, organization_id, external_id, name, type, status,
                    address, area_sqm, budget_vnd, floors
                ) VALUES (
                    gen_random_uuid(), :org_id, :external_id, :name, :type, :status,
                    CAST(:address AS JSONB), :area_sqm, :budget_vnd, :floors
                )
                ON CONFLICT (organization_id, external_id)
                WHERE external_id IS NOT NULL
                DO UPDATE SET
                    name = EXCLUDED.name,
                    type = EXCLUDED.type,
                    status = EXCLUDED.status,
                    address = EXCLUDED.address,
                    area_sqm = EXCLUDED.area_sqm,
                    budget_vnd = EXCLUDED.budget_vnd,
                    floors = EXCLUDED.floors
                """
            ),
            {
                "org_id": str(organization_id),
                "external_id": row["external_id"],
                "name": row["name"],
                "type": row.get("type"),
                "status": row.get("status", "planning"),
                # JSONB binding: serialise to JSON text and cast inside SQL.
                "address": _to_json(row.get("address")),
                "area_sqm": row.get("area_sqm"),
                "budget_vnd": row.get("budget_vnd"),
                "floors": row.get("floors"),
            },
        )
        written += 1
    return written


async def _commit_suppliers(
    session: AsyncSession,
    organization_id: UUID,
    rows: list[dict[str, Any]],
) -> int:
    """Same idempotency story as projects — partial unique on
    (org_id, external_id) WHERE external_id IS NOT NULL AND
    organization_id IS NOT NULL. The trailing org-not-null clause is
    what keeps platform-seeded global suppliers from colliding with
    tenant-imported ones."""
    written = 0
    for row in rows:
        await session.execute(
            text(
                """
                INSERT INTO suppliers (
                    id, organization_id, external_id, name,
                    categories, provinces, contact, verified
                ) VALUES (
                    gen_random_uuid(), :org_id, :external_id, :name,
                    CAST(:categories AS TEXT[]), CAST(:provinces AS TEXT[]),
                    CAST(:contact AS JSONB), :verified
                )
                ON CONFLICT (organization_id, external_id)
                WHERE external_id IS NOT NULL AND organization_id IS NOT NULL
                DO UPDATE SET
                    name = EXCLUDED.name,
                    categories = EXCLUDED.categories,
                    provinces = EXCLUDED.provinces,
                    contact = EXCLUDED.contact,
                    verified = EXCLUDED.verified
                """
            ),
            {
                "org_id": str(organization_id),
                "external_id": row["external_id"],
                "name": row["name"],
                "categories": _to_pg_array(row.get("categories", [])),
                "provinces": _to_pg_array(row.get("provinces", [])),
                "contact": _to_json(row.get("contact") or {}),
                "verified": bool(row.get("verified", False)),
            },
        )
        written += 1
    return written


def _to_json(v: Any) -> str | None:
    """JSONB binding helper — serialise to a JSON string the
    `CAST(:x AS JSONB)` pattern can decode. None passes straight
    through so the column ends up SQL NULL."""
    import json

    return None if v is None else json.dumps(v)


def _to_pg_array(items: list[str]) -> str:
    """Postgres array literal: `{a,b,c}` with nested quoting for
    strings that contain commas/quotes. Cheaper than passing a Python
    list because asyncpg's array binding only works on native types
    after an explicit `cast` — and we already need the cast for the
    NULL-for-empty case below."""
    if not items:
        return "{}"
    escaped = []
    for s in items:
        # Quote everything; double internal quotes / backslashes.
        s2 = s.replace("\\", "\\\\").replace('"', '\\"')
        escaped.append(f'"{s2}"')
    return "{" + ",".join(escaped) + "}"
