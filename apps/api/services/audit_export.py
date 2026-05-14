"""KTNN-friendly audit-log export.

Vietnamese State Audit Office (Kiểm toán Nhà nước — KTNN) reviews of
SOE construction projects routinely demand a date-range dump of the
platform's audit_events table for a specific project / org. Auditors
take this offline and grep / pivot it in Excel.

Format requirements (per KTNN inspector feedback on similar SaaS):

  * **CSV**: ISO-8601 timestamps, semantically named columns (no
    internal UUIDs in the header), one row per event. Vietnamese
    column headers — auditors are non-technical end users.
  * **XLSX**: two sheets:
      1. `Audit Trail` — the same data as the CSV.
      2. `Provenance` — a metadata sheet with the org name, the
         caller's email, the date range, total row count, and a
         SHA-256 digest of the CSV body. The hash is the legal-
         admissibility hook: if a print-out is challenged in court,
         the auditor can rerun the export and compare hashes to
         prove the export wasn't tampered with mid-flight.

The query path uses `AdminSessionFactory` because RLS on audit_events
would silently drop rows from other tenants — fine for normal API
traffic, but the auditor's caller is `Role.ADMIN` of their org, so we
explicitly scope the SQL `WHERE organization_id = :org_id` rather than
relying on the RLS guard-rail (defense in depth).

`_MAX_AUDIT_EXPORT_ROWS` caps the response so a multi-year query can't
OOM the worker. Auditors who need more than 250k rows can split the
date range across multiple exports.
"""

from __future__ import annotations

import csv
import hashlib
import io
from datetime import UTC, date, datetime
from typing import Any
from uuid import UUID

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession


_MAX_AUDIT_EXPORT_ROWS = 250_000

# Vietnamese column headers — KTNN auditors typically don't read English.
# The English internal column name is kept as a comment for grep-ability.
_VI_HEADERS = [
    "Thời điểm",        # created_at
    "Hành động",        # action
    "Loại đối tượng",    # resource_type
    "ID đối tượng",      # resource_id (UUID)
    "Người thực hiện",   # actor_email (joined from users)
    "Loại tài khoản",    # actor_kind (user / api_key / system)
    "Địa chỉ IP",        # ip
    "Trình duyệt",       # user_agent
    "Trạng thái trước",  # before (JSONB → JSON string)
    "Trạng thái sau",    # after (JSONB → JSON string)
]


# ---------- Data fetch ----------


async def _fetch_audit_rows(
    *,
    session: AsyncSession,
    organization_id: UUID,
    since: date | datetime,
    until: date | datetime,
    resource_type: str | None,
    cap: int,
) -> list[dict[str, Any]]:
    """Pull audit_events for the date range, left-joined to users for actor email.

    Ordered by `created_at` ASC so the CSV reads top-to-bottom in time
    order, which is how auditors expect to scan a paper trail.
    """
    where_clauses = [
        "a.organization_id = :org_id",
        "a.created_at >= :since",
        "a.created_at < :until",
    ]
    params: dict[str, Any] = {
        "org_id": str(organization_id),
        "since": since,
        "until": until,
        "cap": cap,
    }
    if resource_type:
        where_clauses.append("a.resource_type = :rt")
        params["rt"] = resource_type

    sql = f"""
        SELECT
            a.created_at,
            a.action,
            a.resource_type,
            a.resource_id,
            a.actor_user_id,
            a.actor_api_key_id,
            u.email AS actor_email,
            a.ip,
            a.user_agent,
            a.before,
            a.after
        FROM audit_events a
        LEFT JOIN users u ON u.id = a.actor_user_id
        WHERE {' AND '.join(where_clauses)}
        ORDER BY a.created_at ASC
        LIMIT :cap
    """
    result = await session.execute(text(sql), params)
    return [dict(r) for r in result.mappings().all()]


# ---------- Row → CSV-cell projection ----------


def _actor_kind(row: dict[str, Any]) -> str:
    if row["actor_user_id"] is not None:
        return "user"
    if row["actor_api_key_id"] is not None:
        return "api_key"
    return "system"


def _stringify_json(v: Any) -> str:
    """JSONB cells round-trip through asyncpg as dicts/lists — render
    them as compact JSON for the spreadsheet. Empty dicts become an
    empty cell instead of `{}` so a blank action stays visually quiet.
    """
    import json

    if v is None or v == {} or v == []:
        return ""
    return json.dumps(v, ensure_ascii=False, separators=(",", ":"))


def _row_to_cells(row: dict[str, Any]) -> list[str]:
    """Project one audit row into the 10 KTNN-format columns."""
    created_at = row["created_at"]
    if isinstance(created_at, datetime):
        ts = created_at.astimezone(UTC).isoformat()
    else:
        ts = str(created_at)
    return [
        ts,
        row["action"] or "",
        row["resource_type"] or "",
        str(row["resource_id"]) if row["resource_id"] is not None else "",
        row.get("actor_email") or "",
        _actor_kind(row),
        row.get("ip") or "",
        row.get("user_agent") or "",
        _stringify_json(row.get("before")),
        _stringify_json(row.get("after")),
    ]


# ---------- CSV ----------


def render_csv(rows: list[dict[str, Any]]) -> bytes:
    """Serialise rows to a UTF-8 CSV body.

    Returns bytes so the caller can both stream them and feed them
    into the SHA-256 digest used by the XLSX provenance sheet.
    """
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_VI_HEADERS)
    for row in rows:
        writer.writerow(_row_to_cells(row))
    return buf.getvalue().encode("utf-8")


# ---------- XLSX with provenance sheet ----------


def _xlsx_with_provenance(
    *,
    rows: list[dict[str, Any]],
    csv_bytes: bytes,
    organization_name: str,
    organization_id: UUID,
    requester_email: str,
    since: date | datetime,
    until: date | datetime,
    resource_type: str | None,
) -> bytes:
    """Build a two-sheet XLSX: Audit Trail + Provenance.

    The Provenance sheet's SHA-256 is computed over the CSV body so
    auditors can re-run the same query and verify byte-for-byte that
    the export wasn't altered after generation. Date-range parameters
    are stamped on the provenance sheet so a print-out is self-
    describing without needing the query URL.
    """
    from openpyxl import Workbook  # heavy import; lazy.

    digest = hashlib.sha256(csv_bytes).hexdigest()

    wb = Workbook()
    # The default "Sheet" is renamed; using `create_sheet` then deleting
    # the default produces an "openpyxl: empty workbook" warning in some
    # consumers, so we mutate the auto-created sheet directly.
    audit_ws = wb.active
    audit_ws.title = "Audit Trail"
    audit_ws.append(_VI_HEADERS)
    for row in rows:
        audit_ws.append(_row_to_cells(row))

    prov_ws = wb.create_sheet(title="Provenance")
    prov_ws.append(["Field", "Value"])
    prov_ws.append(["Organization", organization_name])
    prov_ws.append(["Organization ID", str(organization_id)])
    prov_ws.append(["Requested by", requester_email])
    prov_ws.append(
        ["Date range (from, inclusive)", _isoformat(since)]
    )
    prov_ws.append(
        ["Date range (to, exclusive)", _isoformat(until)]
    )
    prov_ws.append(["Resource type filter", resource_type or "(all)"])
    prov_ws.append(["Total rows", len(rows)])
    prov_ws.append(["Generated at (UTC)", datetime.now(UTC).isoformat()])
    prov_ws.append(["CSV SHA-256", digest])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _isoformat(v: date | datetime) -> str:
    if isinstance(v, datetime):
        return v.astimezone(UTC).isoformat()
    return v.isoformat()


# ---------- Public entry points ----------


async def export_csv(
    *,
    session: AsyncSession,
    organization_id: UUID,
    since: date | datetime,
    until: date | datetime,
    resource_type: str | None = None,
) -> tuple[bytes, int]:
    """Return (csv_body_bytes, row_count) for the requested date range."""
    rows = await _fetch_audit_rows(
        session=session,
        organization_id=organization_id,
        since=since,
        until=until,
        resource_type=resource_type,
        cap=_MAX_AUDIT_EXPORT_ROWS,
    )
    body = render_csv(rows)
    return body, len(rows)


async def export_xlsx(
    *,
    session: AsyncSession,
    organization_id: UUID,
    organization_name: str,
    requester_email: str,
    since: date | datetime,
    until: date | datetime,
    resource_type: str | None = None,
) -> tuple[bytes, int]:
    """Return (xlsx_bytes, row_count) for the requested date range.

    The XLSX bundles the audit data + a Provenance sheet with the
    SHA-256 of the equivalent CSV body, so re-running and hashing
    `export_csv` yields a byte-for-byte verifier.
    """
    rows = await _fetch_audit_rows(
        session=session,
        organization_id=organization_id,
        since=since,
        until=until,
        resource_type=resource_type,
        cap=_MAX_AUDIT_EXPORT_ROWS,
    )
    csv_bytes = render_csv(rows)
    xlsx_bytes = _xlsx_with_provenance(
        rows=rows,
        csv_bytes=csv_bytes,
        organization_name=organization_name,
        organization_id=organization_id,
        requester_email=requester_email,
        since=since,
        until=until,
        resource_type=resource_type,
    )
    return xlsx_bytes, len(rows)


def max_export_rows() -> int:
    return _MAX_AUDIT_EXPORT_ROWS
