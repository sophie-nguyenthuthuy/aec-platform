"""Supplier-directory export — symmetric counterpart to the import.

Buyers iterate: import a CSV, edit it in Excel, re-import. To make
that round-trip work, we need to export the directory in the SAME
shape the import expects, with header aliases the import recognises.

This module is the renderer; the router endpoint streams the bytes
with the right `Content-Type` and `Content-Disposition`.

Two formats:

  * `render_suppliers_xlsx(rows)` — openpyxl workbook with a single
    sheet, column widths sized for the data, header bold + frozen.
    Numeric phone columns get `@` text-format so leading-zero
    Vietnamese mobile numbers (e.g. `0901234567`) don't get
    silently coerced to int by Excel on re-open.
  * `render_suppliers_csv(rows)` — UTF-8 with BOM (Excel-friendly),
    comma-delimited.

Both write the SAME header row keys the import's `_NAME_ALIASES`
etc. recognise — `Tên`, `Email`, `Số điện thoại`, `Danh mục`,
`Tỉnh thành` — so a round-trip via either format is lossless.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable

# Canonical export header. Matches the first alias in each tuple of
# `core._NAME_ALIASES` etc. so re-importing the file produced here
# always succeeds without fuzzy matching.
_EXPORT_HEADER = ("Tên", "Email", "Số điện thoại", "Danh mục", "Tỉnh thành")

# How list-shaped columns (categories, provinces) get serialised.
# `;` is preferred over `,` because `,` is also a CSV delimiter and
# would force the whole cell to be quoted.
_LIST_SEPARATOR = "; "


def _row_cells(row: dict) -> tuple[str, str, str, str, str]:
    """Project a Supplier ORM row → cells in `_EXPORT_HEADER` order.

    `row` is duck-typed: a dict, SQLAlchemy mapping, or an ORM
    instance — anything that exposes `name`, `contact`, `categories`,
    `provinces` as attributes/keys.
    """

    def get(key: str, default=None):
        if isinstance(row, dict):
            return row.get(key, default)
        return getattr(row, key, default)

    name = get("name") or ""
    contact = get("contact") or {}
    email = (contact.get("email") if isinstance(contact, dict) else "") or ""
    phone = (contact.get("phone") if isinstance(contact, dict) else "") or ""
    categories = _LIST_SEPARATOR.join(get("categories") or [])
    provinces = _LIST_SEPARATOR.join(get("provinces") or [])
    return (name, email, phone, categories, provinces)


def render_suppliers_xlsx(rows: Iterable) -> bytes:
    """Render the supplier directory as an .xlsx file.

    Lazy-imports openpyxl so test environments without the dep can
    still load the package.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Suppliers"

    # Header row, bold + light-grey fill for visual separation.
    sheet.append(list(_EXPORT_HEADER))
    header_row = sheet[1]
    bold = Font(bold=True)
    fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    for cell in header_row:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")

    # Body. We iterate `rows` once; if the caller passed an ORM
    # query, that's a one-shot iterable, hence the materialise to
    # tuple here for predictability across calls.
    body_rows = [_row_cells(r) for r in rows]
    for r in body_rows:
        sheet.append(list(r))

    # Phone column ('C') as text so leading-zero VN mobile numbers
    # round-trip without becoming int.
    for cell in sheet["C"][1:]:
        cell.number_format = "@"

    # Reasonable column widths: name + categories are the wide ones.
    widths = {"A": 32, "B": 30, "C": 18, "D": 28, "E": 24}
    for col, w in widths.items():
        sheet.column_dimensions[col].width = w

    # Freeze the header row so scrolling stays oriented.
    sheet.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_suppliers_csv(rows: Iterable) -> bytes:
    """Render as UTF-8-with-BOM CSV — opens cleanly in Excel.

    `csv` module is stdlib so no lazy import.
    """
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(list(_EXPORT_HEADER))
    for r in rows:
        writer.writerow(_row_cells(r))
    # `utf-8-sig` prepends ﻿ so Excel auto-detects UTF-8.
    return out.getvalue().encode("utf-8-sig")
