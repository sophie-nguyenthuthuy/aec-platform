"""openpyxl adapter for supplier import.

Lazy-imports openpyxl so unit-testing the core column-detection logic
doesn't require the C-extension dep. Reads only the active sheet — if
the buyer's workbook has multiple sheets, the first one wins (which is
what openpyxl's `wb.active` returns).
"""

from __future__ import annotations

from io import BytesIO

from .core import SupplierImportError, SupplierRow, coerce_row, detect_columns


def parse_suppliers_xlsx(content: bytes) -> list[SupplierRow]:
    """Parse a .xlsx upload into supplier rows.

    Strategy:
      1. Open the workbook (read-only mode — no formula evaluation,
         no chart loading; ~10x faster on big files and avoids
         openpyxl's unevaluated-formula gotchas).
      2. Use the first non-blank row as the header.
      3. Body rows below it; blank rows are skipped silently.

    Raises:
      * `SupplierImportError` if openpyxl can't open the bytes (looks
        like a non-xlsx upload) or the header doesn't have a name
        column.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover — deployed env always has it
        raise SupplierImportError("openpyxl not installed; cannot parse .xlsx uploads") from exc

    try:
        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise SupplierImportError(f"could not open .xlsx content: {exc}") from exc

    sheet = wb.active
    if sheet is None:  # pragma: no cover — empty workbook is rare
        return []

    rows_iter = sheet.iter_rows(values_only=True)
    header: list[str] | None = None
    bodies: list[list[str]] = []
    for raw in rows_iter:
        # Skip leading blank rows so a spreadsheet with a title banner
        # at the top still parses (the header is the first non-empty
        # row, not literally row 1).
        cells = [_cell_to_str(c) for c in raw]
        if not any(cells):
            continue
        if header is None:
            header = cells
        else:
            bodies.append(cells)

    if header is None:
        raise SupplierImportError("workbook contains no rows")

    cols = detect_columns(header)
    out: list[SupplierRow] = []
    for body in bodies:
        row = coerce_row(body, cols)
        if row is not None:
            out.append(row)
    return out


def _cell_to_str(value) -> str:
    """openpyxl returns native types (int / datetime / None / str). The
    column-detection layer expects strings; coerce here once."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)
