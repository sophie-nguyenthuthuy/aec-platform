"""Tests for the data-export pipeline.

Three layers:

  * Pure helpers — `build_select`, `_csv_cell`, the column allowlist.
    No session.

  * Router happy path — GET /api/v1/export/{entity} returns a 200
    with the expected Content-Type / Content-Disposition headers and
    streams CSV bytes.

  * Format + filter wiring — XLSX path produces a valid xlsx body
    (round-trip via openpyxl), bad UUIDs in `project_id` flip a 400.

Tenant-isolation is exercised at the SQL-shape layer (`build_select`
adds the org_id WHERE) — running against a real cluster would
duplicate the broader RLS test coverage in test_rls.py.
"""

from __future__ import annotations

import io
from collections.abc import AsyncIterator
from contextlib import asynccontextmanager
from datetime import UTC, date, datetime
from typing import Any
from unittest.mock import MagicMock
from uuid import UUID, uuid4

import pytest
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient

from middleware.auth import AuthContext, require_auth  # noqa: F401
from middleware.rbac import Role, require_min_role
from services.exports import (
    EXPORT_CONFIGS,
    _csv_cell,
    build_select,
    write_csv_lines,
)

pytestmark = pytest.mark.asyncio


ORG_ID = UUID("77777777-7777-7777-7777-777777777777")
USER_ID = UUID("88888888-8888-8888-8888-888888888888")


# ---------- FakeAsyncSession ----------


class FakeAsyncSession:
    def __init__(self) -> None:
        self.calls: list[tuple[Any, dict[str, Any]]] = []
        self._results: list[Any] = []

    def push(self, result: Any) -> None:
        self._results.append(result)

    async def __aenter__(self):
        return self

    async def __aexit__(self, *exc):
        return None

    async def commit(self) -> None: ...
    async def close(self) -> None: ...

    async def execute(self, stmt: Any, params: dict[str, Any] | None = None) -> Any:
        self.calls.append((stmt, params or {}))
        if self._results:
            return self._results.pop(0)
        r = MagicMock()
        r.mappings.return_value = []
        return r


@pytest.fixture
def fake_db() -> FakeAsyncSession:
    return FakeAsyncSession()


@pytest.fixture(autouse=True)
def patch_tenant_session(fake_db, monkeypatch):
    """Replace TenantAwareSession with a CM yielding the shared fake.
    Same pattern as test_search_router / test_imports."""

    @asynccontextmanager
    async def _factory(_org_id: UUID) -> AsyncIterator[FakeAsyncSession]:
        yield fake_db

    monkeypatch.setattr("routers.exports.TenantAwareSession", _factory)
    yield fake_db


# ---------- Pure-helper tests ----------


async def test_csv_cell_handles_typed_values():
    """Datetimes ISO; UUID hex; None → empty; bool → lowercase. Pin the
    contract because the import path's validators read these back."""
    assert _csv_cell(None) == ""
    assert _csv_cell(True) == "true"
    assert _csv_cell(False) == "false"
    assert _csv_cell(datetime(2026, 5, 4, 9, 30, tzinfo=UTC)) == "2026-05-04T09:30:00+00:00"
    assert _csv_cell(date(2026, 5, 4)) == "2026-05-04"
    u = UUID("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa")
    assert _csv_cell(u) == str(u)
    assert _csv_cell("plain") == "plain"


async def test_build_select_appends_org_id_filter():
    """`organization_id = :org_id` is mandatory — never a `SELECT`
    without it. Pin so a refactor that loses the WHERE clause breaks
    this test before it ships data exfiltration."""
    config = EXPORT_CONFIGS["projects"]
    sql, _ = build_select(config=config, filters={})
    assert "p.organization_id = :org_id" in sql
    assert sql.startswith("SELECT")


async def test_build_select_drops_unknown_filters_silently():
    """Filter keys outside the per-entity allowlist must be ignored —
    `build_select` is the second line of defense after the router's
    Query() declaration. Pass a junk filter; assert it never reaches
    the bound params."""
    config = EXPORT_CONFIGS["projects"]
    sql, params = build_select(
        config=config,
        filters={"status": "construction", "DROP TABLE": "x"},
    )
    assert "DROP TABLE" not in sql
    assert "status" in params
    assert "DROP TABLE" not in params


async def test_build_select_coerces_uuid_filter():
    """A string `project_id` should round-trip to a UUID instance for
    asyncpg. Bad UUID raises ValueError — router maps to 400."""
    config = EXPORT_CONFIGS["defects"]
    pid = uuid4()
    _, params = build_select(
        config=config,
        filters={"project_id": str(pid)},
    )
    assert params["project_id"] == pid


async def test_build_select_rejects_malformed_uuid():
    config = EXPORT_CONFIGS["defects"]
    with pytest.raises(ValueError, match="invalid project_id"):
        build_select(config=config, filters={"project_id": "not-a-uuid"})


async def test_write_csv_lines_emits_header_first():
    """Header row first; then one chunk per data row. Quoting honours
    csv.QUOTE_MINIMAL so a value with a comma gets wrapped without
    breaking the row count."""
    chunks = list(
        write_csv_lines(
            [{"name": "Tower, A", "n": 1}, {"name": "Tower B", "n": 2}],
            headers=["name", "n"],
        )
    )
    assert len(chunks) == 3  # header + 2 rows
    assert chunks[0].decode().startswith("name,n")
    assert b'"Tower, A"' in chunks[1]


# ---------- Router happy path ----------


def _build_app(role: str = "admin") -> FastAPI:
    from fastapi import HTTPException

    from core.envelope import http_exception_handler, unhandled_exception_handler
    from routers import exports as exports_router

    app = FastAPI()
    app.add_exception_handler(HTTPException, http_exception_handler)
    app.add_exception_handler(Exception, unhandled_exception_handler)
    app.include_router(exports_router.router)

    auth_ctx = AuthContext(
        user_id=USER_ID,
        organization_id=ORG_ID,
        role=role,
        email="caller@example.com",
    )
    app.dependency_overrides[require_min_role(Role.ADMIN)] = lambda: auth_ctx
    app.dependency_overrides[require_auth] = lambda: auth_ctx
    return app


async def test_export_projects_csv_streams_with_attachment_disposition(fake_db):
    """End-to-end: GET /export/projects returns a CSV with the right
    headers, body starts with the column header line, and rows
    follow."""
    rows = MagicMock()
    rows.mappings.return_value = [
        {
            "id": uuid4(),
            "external_id": "P-1",
            "name": "Tower A",
            "type": "office",
            "status": "construction",
            "city": "Hà Nội",
            "district": "Cầu Giấy",
            "area_sqm": 1200.0,
            "budget_vnd": 5_000_000_000,
            "floors": 12,
            "start_date": date(2026, 1, 1),
            "end_date": None,
            "created_at": datetime(2026, 4, 1, tzinfo=UTC),
        }
    ]
    fake_db.push(rows)

    app = _build_app()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        res = await ac.get("/api/v1/export/projects?status=construction")
    assert res.status_code == 200, res.text
    assert "text/csv" in res.headers["content-type"]
    assert "attachment;" in res.headers["content-disposition"]
    assert "aec-projects-" in res.headers["content-disposition"]
    body = res.text
    # First line is the header — pin the exact column ordering.
    first_line = body.splitlines()[0]
    assert first_line.startswith(
        "id,external_id,name,type,status,city,district,area_sqm,budget_vnd,floors,start_date,end_date,created_at"
    )
    # Vietnamese chars survive the round-trip (UTF-8).
    assert "Hà Nội" in body
    assert "Tower A" in body
    # Bound params: org_id present, status filter coerced to bound name.
    params = fake_db.calls[0][1]
    assert params["org_id"] == str(ORG_ID)
    assert params["status"] == "construction"


async def test_export_xlsx_returns_valid_workbook(fake_db):
    """XLSX path: response body must be a real openpyxl-readable .xlsx.
    Round-trip through `load_workbook` to verify the bytes aren't a
    truncated CSV or an HTML error page."""
    rows = MagicMock()
    rows.mappings.return_value = [
        {
            "id": uuid4(),
            "external_id": "S-1",
            "name": "Acme Cement",
            "categories": "cement, concrete",
            "provinces": "HN",
            "phone": "+84 123",
            "email": None,
            "address": None,
            "verified": True,
            "rating": None,
            "created_at": datetime(2026, 4, 1, tzinfo=UTC),
        }
    ]
    fake_db.push(rows)

    app = _build_app()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        res = await ac.get("/api/v1/export/suppliers?format=xlsx&verified=true")
    assert res.status_code == 200
    assert "spreadsheetml.sheet" in res.headers["content-type"]
    # Bind path: verified=true coerced to a bool for the SQL filter.
    params = fake_db.calls[0][1]
    assert params["verified"] is True

    # Round-trip through openpyxl.
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(res.content), read_only=True)
    ws = wb.active
    rows_read = list(ws.iter_rows(values_only=True))
    # First row is the header.
    assert rows_read[0][0] == "id"
    # Data row matches what we pushed.
    assert "Acme Cement" in rows_read[1]


async def test_export_threads_filters_through_to_sql(fake_db):
    """Multiple filters → multiple bound params + matching SQL
    fragments. Pin the project_id filter especially; it goes through
    UUID coercion."""
    pid = uuid4()
    rows = MagicMock()
    rows.mappings.return_value = []
    fake_db.push(rows)

    app = _build_app()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        res = await ac.get(f"/api/v1/export/defects?project_id={pid}&status=open&priority=high&since=2026-01-01")
    assert res.status_code == 200
    sql = str(fake_db.calls[0][0])
    params = fake_db.calls[0][1]
    assert "d.project_id = :project_id" in sql
    assert "d.status = :status" in sql
    assert "d.priority = :priority" in sql
    assert "d.reported_at >= :since" in sql
    assert params["project_id"] == pid
    assert params["since"] == date(2026, 1, 1)


async def test_export_xlsx_returns_400_on_bad_filter(fake_db):
    """Bad UUID in `project_id` for XLSX: the build runs eagerly so we
    can still 400 before committing the response. CSV path is more
    forgiving (already streaming) — exercised in a separate test."""
    app = _build_app()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        res = await ac.get("/api/v1/export/defects?format=xlsx&project_id=not-a-uuid")
    assert res.status_code == 400


async def test_export_csv_emits_inline_error_marker_on_bad_filter(fake_db):
    """CSV path can't 400 (response body already committed by the
    first yield), so it emits a `# error: …` line and ends. Verify the
    marker lands so a customer's pipeline can detect failure."""
    app = _build_app()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        res = await ac.get("/api/v1/export/defects?project_id=not-a-uuid")
    # The streaming CSV returns 200, body has the error marker.
    assert res.status_code == 200
    assert res.text.startswith("# error:")


# ---------- Auth ----------


async def test_export_rejects_non_admin():
    """Members shouldn't bulk-download the org's data — the page UI
    already shows them most of it, but the audit story for who pulled
    a CSV is admin-only."""
    from fastapi import HTTPException

    from core.envelope import http_exception_handler, unhandled_exception_handler
    from routers import exports as exports_router

    app = FastAPI()
    app.add_exception_handler(HTTPException, http_exception_handler)
    app.add_exception_handler(Exception, unhandled_exception_handler)
    app.include_router(exports_router.router)
    member_ctx = AuthContext(
        user_id=USER_ID,
        organization_id=ORG_ID,
        role="member",
        email="m@example.com",
    )
    app.dependency_overrides[require_auth] = lambda: member_ctx

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        res = await ac.get("/api/v1/export/projects")
    assert res.status_code == 403


async def test_export_unknown_entity_is_422():
    """The Literal on the path param keeps junk slugs out — pin so a
    refactor that loosens the type doesn't expose `SELECT *` against
    an arbitrary FROM clause."""
    app = _build_app()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        res = await ac.get("/api/v1/export/proposals")
    assert res.status_code == 422
