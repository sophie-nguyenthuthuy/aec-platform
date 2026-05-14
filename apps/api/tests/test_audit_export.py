"""Unit tests for services.audit_export pure helpers.

The DB-touching code paths (export_csv, export_xlsx) are covered by
the live `tests/test_vn_modules_integration.py` suite that runs against
Postgres. Here we cover the pure functions that need no live DB:

  * `_row_to_cells` produces the right 10-column projection
  * `render_csv` emits a valid CSV with Vietnamese headers
  * `_xlsx_with_provenance` produces a workbook with two sheets, the
    second of which contains a SHA-256 of the CSV body that's
    re-derivable by hashing `render_csv` directly.

The legal-admissibility hook (re-running and comparing hashes) is the
load-bearing test here — if the digest computation drifts between the
CSV path and the XLSX provenance path, KTNN auditors lose the ability
to prove an exported file wasn't tampered with.
"""

from __future__ import annotations

import csv
import hashlib
import io
from datetime import UTC, datetime
from uuid import uuid4


def test_row_to_cells_user_kind():
    """Row with `actor_user_id` set → actor_kind = 'user'."""
    from services.audit_export import _row_to_cells

    row = {
        "created_at": datetime(2026, 1, 15, 10, 30, tzinfo=UTC),
        "action": "project.create",
        "resource_type": "projects",
        "resource_id": uuid4(),
        "actor_user_id": uuid4(),
        "actor_api_key_id": None,
        "actor_email": "alice@example.com",
        "ip": "10.0.0.1",
        "user_agent": "Mozilla",
        "before": {},
        "after": {"name": "Khu chung cư"},
    }
    cells = _row_to_cells(row)
    assert len(cells) == 10
    assert cells[0].startswith("2026-01-15T10:30")
    assert cells[1] == "project.create"
    assert cells[2] == "projects"
    assert cells[4] == "alice@example.com"
    assert cells[5] == "user"
    # JSON serialisation must round-trip Vietnamese names without escaping
    assert "Khu chung cư" in cells[9]


def test_row_to_cells_system_kind():
    """Row with both actor fields NULL → actor_kind = 'system'."""
    from services.audit_export import _row_to_cells

    row = {
        "created_at": datetime(2026, 1, 1, tzinfo=UTC),
        "action": "cron.tick",
        "resource_type": "cron",
        "resource_id": None,
        "actor_user_id": None,
        "actor_api_key_id": None,
        "actor_email": None,
        "ip": None,
        "user_agent": None,
        "before": None,
        "after": None,
    }
    cells = _row_to_cells(row)
    assert cells[5] == "system"
    # Empty JSON dicts/None must render as empty string, not "{}" or "null"
    assert cells[8] == ""
    assert cells[9] == ""


def test_row_to_cells_api_key_kind():
    """Row with only `actor_api_key_id` → actor_kind = 'api_key'."""
    from services.audit_export import _row_to_cells

    row = {
        "created_at": datetime(2026, 2, 1, tzinfo=UTC),
        "action": "rfq.create",
        "resource_type": "rfqs",
        "resource_id": uuid4(),
        "actor_user_id": None,
        "actor_api_key_id": uuid4(),
        "actor_email": None,
        "ip": "192.0.2.1",
        "user_agent": "aec-cli/1.0",
        "before": {},
        "after": {},
    }
    cells = _row_to_cells(row)
    assert cells[5] == "api_key"


def test_render_csv_has_vietnamese_headers():
    """First row must be the 10 Vietnamese column headers."""
    from services.audit_export import render_csv, _VI_HEADERS

    body = render_csv([])
    rows = list(csv.reader(io.StringIO(body.decode("utf-8"))))
    assert rows[0] == _VI_HEADERS


def test_xlsx_provenance_hash_matches_csv():
    """The XLSX Provenance sheet's SHA-256 must match the bare CSV's SHA-256.

    This is the load-bearing test for the legal-admissibility story:
    auditors can re-run the CSV export and verify the hash on the
    Provenance sheet matches, proving no tampering.
    """
    from openpyxl import load_workbook

    from services.audit_export import _xlsx_with_provenance, render_csv

    rows = [
        {
            "created_at": datetime(2026, 1, 1, 12, 0, tzinfo=UTC),
            "action": "test.event",
            "resource_type": "test",
            "resource_id": None,
            "actor_user_id": None,
            "actor_api_key_id": None,
            "actor_email": None,
            "ip": None,
            "user_agent": None,
            "before": {},
            "after": {"k": "v"},
        }
    ]
    csv_bytes = render_csv(rows)
    expected_digest = hashlib.sha256(csv_bytes).hexdigest()

    xlsx_bytes = _xlsx_with_provenance(
        rows=rows,
        csv_bytes=csv_bytes,
        organization_name="Cty Xây dựng TNHH ABC",
        organization_id=uuid4(),
        requester_email="auditor@example.com",
        since=datetime(2026, 1, 1, tzinfo=UTC),
        until=datetime(2026, 2, 1, tzinfo=UTC),
        resource_type=None,
    )

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == ["Audit Trail", "Provenance"]

    prov = wb["Provenance"]
    # Walk the key/value rows looking for the digest cell.
    digest_in_sheet = None
    for row in prov.iter_rows(values_only=True):
        if row[0] == "CSV SHA-256":
            digest_in_sheet = row[1]
            break
    assert digest_in_sheet == expected_digest, (
        "Provenance sheet hash drifted from raw CSV hash — KTNN tamper-detection broken"
    )

    # And the Vietnamese org name must round-trip through XLSX correctly.
    org_row = next(r for r in prov.iter_rows(values_only=True) if r[0] == "Organization")
    assert org_row[1] == "Cty Xây dựng TNHH ABC"


def test_render_csv_escapes_commas_in_vietnamese_text():
    """Vietnamese supplier names with embedded commas must round-trip
    through csv.reader without splitting incorrectly."""
    from services.audit_export import render_csv

    rows = [
        {
            "created_at": datetime(2026, 1, 1, tzinfo=UTC),
            "action": "supplier.update",
            "resource_type": "suppliers",
            "resource_id": uuid4(),
            "actor_user_id": uuid4(),
            "actor_api_key_id": None,
            "actor_email": "ops@x.vn",
            "ip": None,
            "user_agent": None,
            "before": {"name": "Vật liệu, Phụ kiện Hòa Phát"},
            "after": {"name": "Vật liệu, Phụ kiện Hòa Phát LTD"},
        }
    ]
    body = render_csv(rows)
    parsed = list(csv.reader(io.StringIO(body.decode("utf-8"))))
    # Header + 1 data row.
    assert len(parsed) == 2
    # The `after` cell is the last column — should still be one cell
    # despite the comma.
    assert "Vật liệu, Phụ kiện Hòa Phát LTD" in parsed[1][-1]
