"""Unit tests for `services.boq_io`.

Three layers under test:

  * `core` — column detection, decimal coercion, row-to-grid mapping.
    Pure list-of-list logic, no Excel/PDF deps required.
  * `xlsx` — openpyxl roundtrip: render → parse must reproduce input.
  * `pdf`  — reportlab smoke render — assert non-empty bytes + valid
    PDF magic. We don't OCR the output; that's reportlab's job.

Router-level integration sits in `test_costpulse_router.py` once the
import/export endpoints land.
"""

from __future__ import annotations

from decimal import Decimal

import pytest

from services.boq_io import (
    BoqIOError,
    BoqRow,
    detect_columns,
    parse_boq_xlsx,
    render_boq_pdf,
    render_boq_xlsx,
    rows_to_grid,
)
from services.boq_io.core import coerce_decimal, coerce_str

# ---------- Column detection ----------


class TestDetectColumns:
    def test_canonical_vietnamese_headers(self):
        cols = detect_columns(["STT", "Mô tả công việc", "Đơn vị", "Khối lượng", "Đơn giá", "Thành tiền"])
        assert cols is not None
        assert cols.code == 0
        assert cols.description == 1
        assert cols.unit == 2
        assert cols.quantity == 3
        assert cols.unit_price_vnd == 4
        assert cols.total_price_vnd == 5

    def test_english_headers(self):
        cols = detect_columns(["No.", "Description", "Unit", "Qty", "Unit Price", "Total"])
        assert cols is not None
        assert cols.description == 1
        assert cols.quantity == 3
        assert cols.unit_price_vnd == 4
        assert cols.total_price_vnd == 5

    def test_diacritics_folded_for_matching(self):
        # Source spreadsheet has no Vietnamese accents — must still match.
        cols = detect_columns(["Mo ta", "Don vi", "Don gia"])
        assert cols is not None
        assert cols.description == 0
        assert cols.unit == 1
        assert cols.unit_price_vnd == 2

    def test_returns_none_when_description_missing(self):
        # Without a description column we can't import anything.
        assert detect_columns(["STT", "Đơn vị", "Đơn giá"]) is None

    def test_quantity_column_optional(self):
        cols = detect_columns(["Description", "Unit Price"])
        assert cols is not None
        assert cols.description == 0
        assert cols.unit_price_vnd == 1
        assert cols.quantity is None

    def test_excludes_already_claimed_index_for_description(self):
        # "Mô tả" matches description; "Tên hàng" would also be a name
        # alias but should NOT shadow column 0.
        cols = detect_columns(["Mô tả", "Đơn vị", "Đơn giá"])
        assert cols is not None
        assert cols.description == 0


# ---------- Decimal coercion ----------


class TestCoerceDecimal:
    @pytest.mark.parametrize(
        "raw,expected",
        [
            (1234, Decimal("1234")),
            (1234.5, Decimal("1234.5")),
            ("1,234,567", Decimal("1234567")),
            ("1.234.567", Decimal("1234567")),  # Vietnamese thousands
            ("1.234.567,89", Decimal("1234567.89")),  # Vietnamese decimal
            ("12,500", Decimal("12500")),  # 3-digit grouping → thousands
            ("1,5", Decimal("1.5")),  # short comma → decimal
            ("2 000 000 đ", Decimal("2000000")),
            ("VND 1,234,567.50", Decimal("1234567.50")),
            (None, None),
            ("", None),
            ("—", None),
            ("N/A", None),
        ],
    )
    def test_handles_common_formats(self, raw, expected):
        assert coerce_decimal(raw) == expected


def test_coerce_str_strips_and_handles_none():
    assert coerce_str("  hello  ") == "hello"
    assert coerce_str("") is None
    assert coerce_str(None) is None
    assert coerce_str(42) == "42"


# ---------- rows_to_grid ----------


def test_rows_to_grid_uses_canonical_column_order():
    rows = [
        BoqRow(
            description="Bê tông C30",
            code="1.1",
            unit="m3",
            quantity=Decimal("120"),
            unit_price_vnd=Decimal("2050000"),
            material_code="CONC_C30",
        ),
    ]
    header, body = rows_to_grid(rows)
    assert header == [
        "Code",
        "Description",
        "Unit",
        "Quantity",
        "Unit price (VND)",
        "Total (VND)",
        "Material code",
    ]
    # Total auto-computed from quantity × unit_price when missing.
    assert body[0] == ["1.1", "Bê tông C30", "m3", 120.0, 2050000.0, 246000000.0, "CONC_C30"]


def test_rows_to_grid_preserves_explicit_total_over_computed():
    rows = [
        BoqRow(
            description="Item with quoted total",
            quantity=Decimal("100"),
            unit_price_vnd=Decimal("1000"),
            total_price_vnd=Decimal("99500"),  # supplier discount → don't recompute
        ),
    ]
    _, body = rows_to_grid(rows)
    assert body[0][5] == 99500.0


def test_rows_to_grid_emits_empty_strings_for_missing_values():
    rows = [BoqRow(description="Just a description")]
    _, body = rows_to_grid(rows)
    assert body[0] == ["", "Just a description", "", "", "", "", ""]


# ---------- xlsx roundtrip ----------


class TestXlsxRoundtrip:
    def test_render_then_parse_reproduces_rows(self):
        original = [
            BoqRow(
                description="Bê tông thương phẩm C30",
                code="1.1",
                unit="m3",
                quantity=Decimal("120"),
                unit_price_vnd=Decimal("2050000"),
                material_code="CONC_C30",
            ),
            BoqRow(
                description="Thép cuộn CB500",
                code="1.2",
                unit="kg",
                quantity=Decimal("8500"),
                unit_price_vnd=Decimal("20500"),
            ),
        ]
        blob = render_boq_xlsx(original)
        # `bytes` so we can hand to FastAPI.StreamingResponse without
        # extra wrapping. Non-empty + xlsx magic prefix.
        assert isinstance(blob, bytes)
        assert blob[:4] == b"PK\x03\x04"  # zip magic — xlsx is a zip file

        parsed = parse_boq_xlsx(blob)
        assert len(parsed) == 2
        assert parsed[0].description == "Bê tông thương phẩm C30"
        assert parsed[0].code == "1.1"
        assert parsed[0].unit == "m3"
        assert parsed[0].quantity == Decimal("120")
        assert parsed[0].unit_price_vnd == Decimal("2050000")
        assert parsed[0].total_price_vnd == Decimal("246000000")
        assert parsed[0].material_code == "CONC_C30"

    def test_parse_skips_blank_description_rows(self):
        # Hand-build a workbook with title row + header + a blank
        # subtotal row in the middle.
        import openpyxl

        wb = openpyxl.Workbook()
        s = wb.active
        s.append(["Project: Tower X — BOQ"])  # title banner
        s.append([])  # blank
        s.append(["STT", "Mô tả", "Đơn vị", "Khối lượng", "Đơn giá", "Thành tiền"])
        s.append(["1.1", "Bê tông C30", "m3", 120, 2050000, 246000000])
        s.append([None, None, None, None, None, None])  # blank
        s.append(
            [None, "(subtotal)", None, None, None, 246000000]
        )  # ignored by description-required gate? — actually has desc
        s.append(["1.2", "Thép CB500", "kg", 8500, 20500, 174250000])
        import io as io_mod

        buf = io_mod.BytesIO()
        wb.save(buf)
        rows = parse_boq_xlsx(buf.getvalue())
        # All description-bearing rows survive (including the subtotal-
        # like one). The user's spreadsheet is the source of truth; we
        # only drop rows with no description at all.
        descriptions = [r.description for r in rows]
        assert "Bê tông C30" in descriptions
        assert "Thép CB500" in descriptions

    def test_parse_raises_when_no_header_row_found(self):
        import openpyxl

        wb = openpyxl.Workbook()
        s = wb.active
        for _ in range(5):
            s.append(["random", "data", "no", "headers"])
        import io as io_mod

        buf = io_mod.BytesIO()
        wb.save(buf)
        with pytest.raises(BoqIOError, match="no recognisable BOQ header"):
            parse_boq_xlsx(buf.getvalue())

    def test_parse_raises_on_corrupt_bytes(self):
        with pytest.raises(BoqIOError, match="could not open"):
            parse_boq_xlsx(b"not an xlsx file at all")

    def test_render_includes_freeze_pane_and_widths(self):
        """Light sanity check on the styling — full visual review is manual."""
        import io as io_mod

        import openpyxl

        rows = [BoqRow(description="x")]
        blob = render_boq_xlsx(rows)
        wb = openpyxl.load_workbook(io_mod.BytesIO(blob))
        s = wb.active
        assert s.freeze_panes == "A2"
        assert s.column_dimensions["B"].width == 50  # description column
        # Header row exists and is bolded.
        assert s.cell(row=1, column=2).value == "Description"
        assert s.cell(row=1, column=2).font.bold is True


# ---------- pdf smoke ----------


class TestPdfRender:
    def test_renders_valid_pdf_bytes(self):
        rows = [
            BoqRow(
                description="Bê tông C30",
                code="1.1",
                unit="m3",
                quantity=Decimal("120"),
                unit_price_vnd=Decimal("2050000"),
            ),
            BoqRow(description="Description-only row"),
        ]
        blob = render_boq_pdf("Tower X — Schematic v1", rows)
        assert isinstance(blob, bytes)
        # PDF magic: every PDF starts with `%PDF-`.
        assert blob.startswith(b"%PDF-")
        # Output isn't trivially empty.
        assert len(blob) > 1000
        # Contains the estimate name in the rendered content stream
        # (reportlab leaves it readable in the producer-info block).
        assert b"Tower X" in blob or b"BOQ" in blob

    def test_renders_with_zero_rows(self):
        """Empty BOQs shouldn't crash — render the headers + zero body."""
        blob = render_boq_pdf("Empty estimate", [])
        assert blob.startswith(b"%PDF-")

    def test_renders_with_unicode_estimate_name(self):
        """Vietnamese estimate names must round-trip without UnicodeEncodeError."""
        blob = render_boq_pdf("Tòa nhà chung cư", [BoqRow(description="x")])
        assert blob.startswith(b"%PDF-")
